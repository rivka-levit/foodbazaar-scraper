import openpyxl

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Useful for handling different item types with a single interface
from itemadapter import ItemAdapter
from scrapy.exceptions import DropItem

from datetime import datetime as dt

from dotenv import load_dotenv

from foodbazaar.items import FoodbazaarItem

import mysql.connector
import os

load_dotenv()



class RemoveDuplicatesPipeline:
    """Remove duplicated items."""

    def __init__(self):
        self.name_price = set()

    def process_item(self, item: FoodbazaarItem, spider):
        adapter = ItemAdapter(item)

        if (adapter['name'], adapter['price']) in self.name_price:
            raise DropItem(f'Duplicated item found: {item}')

        self.name_price.add((adapter['name'], adapter['price']))
        return item


class AddToXlsxPipeline:
    """Add items to an Excel file."""

    def __init__(self):
        self.workbook = None
        self.worksheet = None
        self.current_row_index = 0

    def process_item(self, item: FoodbazaarItem, spider):

        values = list(item.values())
        letters = [f'A{self.current_row_index + 1}',
                   f'B{self.current_row_index + 1}',
                   f'C{self.current_row_index + 1}']

        # Add header at first row
        if self.current_row_index == 0:

            # Get header from item keys
            header = list(map(lambda x: x.title(), item.keys()))
            self.worksheet.append(header)

            # Add style to header
            self.add_header_style(letters)

            # Add value below the header
            self.worksheet.append(values)
        else:
            self.worksheet.append(values)
            self.add_table_borders(letters)

        self.current_row_index += 1

        return item

    def open_spider(self, spider):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

        # Set row height for the header
        row = self.worksheet.row_dimensions[1]
        row.height = 20

        # Set width for all the columns
        self.worksheet.column_dimensions['A'].width = 30
        self.worksheet.column_dimensions['B'].width = 30
        self.worksheet.column_dimensions['C'].width = 70

        self.conn = mysql.connector.connect(
            host=os.environ.get('DB_HOST'),
            user='root',
            password=os.environ.get('DB_ROOT_PASS'),
            database=os.environ.get('DB_NAME')
        )
        self.cursor = self.conn.cursor()
        # self.cursor.execute(
        #     f"""CREATE DATABASE IF NOT EXISTS {self.database}"""
        # )
        self.cursor.execute(
            f"""CREATE TABLE IF NOT EXISTS foodbazaar (
                        id INT PRIMARY KEY AUTO_INCREMENT,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        filepath VARCHAR(255)
                    )"""
        )

    def close_spider(self, spider):
        max_row = self.worksheet.max_row
        letters = [f'A{max_row}', f'B{max_row}', f'C{max_row}']

        self.add_table_borders(letters)

        filename = f'{str(dt.now().timestamp()).replace(".", "_")}_foodbazaar.xlsx'  # File xlsx name
        filepath = '/vol/uploads/files/' + filename
        self.workbook.save(filepath)

        self.cursor.execute(f"""
                    INSERT INTO {os.environ.get('DB_NAME')}.foodbazaar (filepath)
                    VALUES (%s)
                """, (filepath,))

        self.cursor.execute(f"SELECT id FROM {os.environ.get('DB_NAME')}.foodbazaar WHERE filepath = '/vol/uploads/files/{filename}'")
        print(self.cursor.fetchall()[-1][0])
        self.cursor.close()
        self.conn.close()

    def add_table_borders(self, letters: list[str]) -> None:
        """
        Add border style to the cells in `letters`.

        :param letters: List of cells' names.
        :return: None
        """

        for letter in letters:
            self.worksheet[letter].border = Border(
                top=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'),
                left=Side(border_style='thin', color='000000')
            )

    def add_header_style(self, letters: list[str]) -> None:
        """
        Add style to the cells of head row in the table.

        :param letters: List of cells' names.
        :return: None
        """

        for letter in letters:
            self.worksheet[letter].fill = PatternFill(
                'solid',
                fgColor='dfe4ff'
            )
            self.worksheet[letter].font = Font(
                bold=True,
                name='Arial',
                size=12
            )
            self.worksheet[letter].alignment = Alignment(
                horizontal='center',
                vertical='center'
            )

        self.add_table_borders(letters)
